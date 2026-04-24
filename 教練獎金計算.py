import streamlit as st
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font

def generate_perfect_salary_report(uploaded_files):
    # 教練單價表
    COACH_PRICING = {
        "林意潔": {"團1-2人": 800, "團3人": 900, "團4人": 900, "團5人": 950, "團6人": 1000, "1對2(1.5hr)": 1275, "1對2": 850, "1對1(1.5hr)": 1275, "1對1": 850},
        "陳秀蓉": {"團1-2人": 800, "團3人": 850, "團4人": 850, "團5人": 900, "團6人": 950, "1對2(1.5hr)": 1200, "1對2": 800, "1對1(1.5hr)": 1200, "1對1": 800},
        "陳怡廷": {"團1-2人": 800, "團3人": 850, "團4人": 900, "團5人": 950, "團6人": 1000, "1對2(1.5hr)": 1275, "1對2": 850, "1對1(1.5hr)": 1275, "1對1": 850},
        "鍾佳蓁 Rita": {"團1-2人": 900, "團3人": 950, "團4人": 1000, "團5人": 1050, "團6人": 1100, "1對2(1.5hr)": 1425, "1對2": 950, "1對1(1.5hr)": 1425, "1對1": 950},
        "黃宛婷": {"團1-2人": 700, "團3人": 750, "團4人": 800, "團5人": 850, "團6人": 900, "1對2(1.5hr)": 1125, "1對2": 750, "1對1(1.5hr)": 1125, "1對1": 750},
        "楊子慧(小在)": {"團1-2人": 650, "團3人": 800, "團4人": 800, "團5人": 850, "團6人": 900, "1對2(1.5hr)": 1050, "1對2": 700, "1對1(1.5hr)": 1050, "1對1": 700},
        "許力尹 LOUIS": {"團1-2人": 600, "團3人": 800, "團4人": 800, "團5人": 800, "團6人": 800, "1對2(1.5hr)": 1200, "1對2": 800, "1對1(1.5hr)": 1200, "1對1": 800},
        "顥顥": {"團1-2人": 0, "團3人": 1000, "團4人": 1000, "團5人": 1100, "團6人": 1100, "1對2(1.5hr)": 1500, "1對2": 1000, "1對1(1.5hr)": 1500, "1對1": 1000},
        "洪睿絃": {"團1-2人": 700, "團3人": 750, "團4人": 800, "團5人": 850, "團6人": 900, "1對2(1.5hr)": 1125, "1對2": 750, "1對1(1.5hr)": 1125, "1對1": 750},
        "紀儒蓁": {"團1-2人": 700, "團3人": 800, "團4人": 800, "團5人": 850, "團6人": 900, "1對2(1.5hr)": 1125, "1對2": 750, "1對1(1.5hr)": 1125, "1對1": 750},
        "李翎瑋": {"團1-2人": 800, "團3人": 850, "團4人": 900, "團5人": 950, "團6人": 1000, "1對2(1.5hr)": 1275, "1對2": 850, "1對1(1.5hr)": 1275, "1對1": 850},
        "郭奕伶": {"團1-2人": 550, "團3人": 600, "團4人": 650, "團5人": 700, "團6人": 750, "1對2(1.5hr)": 900, "1對2": 600, "1對1(1.5hr)": 900, "1對1": 600},
        "郭品均": {"團1-2人": 700, "團3人": 750, "團4人": 800, "團5人": 850, "團6人": 900, "1對2(1.5hr)": 1125, "1對2": 750, "1對1(1.5hr)": 1125, "1對1": 750},
        "邴妍語": {"團1-2人": 700, "團3人": 750, "團4人": 800, "團5人": 850, "團6人": 900, "1對2(1.5hr)": 1125, "1對2": 750, "1對1(1.5hr)": 1125, "1對1": 750},
        "張鈞弼": {"團1-2人": 550, "團3人": 600, "團4人": 650, "團5人": 700, "團6人": 850, "1對2(1.5hr)": 900, "1對2": 600, "1對1(1.5hr)": 900, "1對1": 600},
        "蕭竣升": {"團1-2人": 700, "團3人": 750, "團4人": 800, "團5人": 850, "團6人": 900, "1對2(1.5hr)": 1125, "1對2": 750, "1對1(1.5hr)": 1125, "1對1": 750},
        "紀萃文": {"團1-2人": 700, "團3人": 750, "團4人": 800, "團5人": 850, "團6人": 900, "1對2(1.5hr)": 1125, "1對2": 750, "1對1(1.5hr)": 1125, "1對1": 750},
        "李函豫": {"團1-2人": 600, "團3人": 650, "團4人": 700, "團5人": 750, "團6人": 800, "1對2(1.5hr)": 975, "1對2": 650, "1對1(1.5hr)": 975, "1對1": 650},
        "尤子綺": {"團1-2人": 550, "團3人": 600, "團4人": 650, "團5人": 700, "團6人": 750, "1對2(1.5hr)": 900, "1對2": 600, "1對1(1.5hr)": 900, "1對1": 600},
        "張楷翌": {"團1-2人": 600, "團3人": 650, "團4人": 700, "團5人": 750, "團6人": 800, "1對2(1.5hr)": 975, "1對2": 650, "1對1(1.5hr)": 975, "1對1": 650},
        "侯懿庭": {"團1-2人": 700, "團3人": 750, "團4人": 800, "團5人": 850, "團6人": 900, "1對2(1.5hr)": 1125, "1對2": 750, "1對1(1.5hr)": 1125, "1對1": 750},
        "謝俐池": {"團1-2人": 550, "團3人": 600, "團4人": 650, "團5人": 700, "團6人": 750, "1對2(1.5hr)": 900, "1對2": 600, "1對1(1.5hr)": 900, "1對1": 600},
        "黃姿菁": {"團1-2人": 550, "團3人": 600, "團4人": 650, "團5人": 700, "團6人": 750, "1對2(1.5hr)": 900, "1對2": 600, "1對1(1.5hr)": 900, "1對1": 600},
        "籃郁雯": {"團1-2人": 550, "團3人": 600, "團4人": 700, "團5人": 700, "團6人": 700, "1對2(1.5hr)": 900, "1對2": 600, "1對1(1.5hr)": 900, "1對1": 600},
        "徐漫": {"團1-2人": 500, "團3人": 550, "團4人": 600, "團5人": 650, "團6人": 700, "1對2(1.5hr)": 825, "1對2": 550, "1對1(1.5hr)": 825, "1對1": 550},
        "鄭筠馨": {"團1-2人": 550, "團3人": 600, "團4人": 650, "團5人": 700, "團6人": 750, "1對2(1.5hr)": 900, "1對2": 600, "1對1(1.5hr)": 900, "1對1": 600},
        "高舒涵": {"團1-2人": 550, "團3人": 600, "團4人": 650, "團5人": 700, "團6人": 750, "1對2(1.5hr)": 900, "1對2": 600, "1對1(1.5hr)": 900, "1對1": 600},
        "邱靜瑜": {"團1-2人": 600, "團3人": 650, "團4人": 700, "團5人": 750, "團6人": 800, "1對2(1.5hr)": 975, "1對2": 650, "1對1(1.5hr)": 975, "1對1": 650}
    }
    NAME_CONVERSION = {
        "意潔": "林意潔", "Vivi": "陳秀蓉", "怡廷": "陳怡廷", "佳蓁": "鍾佳蓁 Rita", "宛婷": "黃宛婷",
        "小在": "楊子慧(小在)", "LOUIS": "許力尹 LOUIS", "顥顥": "顥顥", "睿絃": "洪睿絃", "儒蓁": "紀儒蓁",
        "翎瑋": "李翎瑋", "奕伶": "郭奕伶", "品均": "郭品均", "妍語": "邴妍語", "鈞弼": "張鈞弼",
        "竣升": "蕭竣升", "萃萃": "紀萃文", "函豫": "李函豫", "子綺": "尤子綺", "楷翌": "張楷翌",
        "懿庭": "侯懿庭", "俐池": "謝俐池", "姿菁": "黃姿菁", "郁雯": "籃郁雯", "徐漫": "徐漫",
        "筠馨": "鄭筠馨", "舒涵": "高舒涵", "靜瑜": "邱靜瑜"
    }
    SENIORITY_DATA = {
        "林意潔": "兩年以上", "陳秀蓉": "兩年以上", "陳怡廷": "兩年以上", "鍾佳蓁 Rita": "兩年以上",
        "黃宛婷": "一年以上至兩年以下", "楊子慧(小在)": "一年以上至兩年以下", "許力尹 LOUIS": "一年以上至兩年以下",
        "顥顥": "一年以上至兩年以下", "洪睿絃": "一年以上至兩年以下", "紀儒蓁": "兩年以上",
        "李翎瑋": "一年以上至兩年以下", "郭奕伶": "一年以上至兩年以下", "郭品均": "一年以上至兩年以下",
        "邴妍語": "一年以上至兩年以下", "張鈞弼": "一年以上至兩年以下", "蕭竣升": "一年以上至兩年以下",
        "紀萃文": "一年以上至兩年以下", "李函豫": "一年以上至兩年以下", "尤子綺": "一年以下",
        "張楷翌": "一年以上至兩年以下", "侯懿庭": "一年以下", "謝俐池": "一年以下",
        "黃姿菁": "一年以下", "籃郁雯": "一年以下", "徐漫": "一年以下",
        "鄭筠馨": "一年以下", "高舒涵": "一年以下", "邱靜瑜": "一年以下"
    }
    
    course_types = ["團1-2人", "團3人", "團4人", "團5人", "團6人", "1對2(1.5hr)", "1對2", "1對1(1.5hr)", "1對1"]
    
    # 定義左右並列的欄位
    columns = [
        "教練_L", "課程_L", "單價_L", "義昌館堂數", "義昌館金額", "高美館堂數", "高美館金額", "中山館堂數", "中山館金額", "三館總堂數", "三館總金額", "應付金額_L", "堂數達標獎金_L", "總計_L", "執行業務_L", "補充保費_L", "應付薪資_L",
        "教練_R", "課程_R", "單價_R", "巨蛋館堂數", "巨蛋館金額", "應付金額_R", "堂數達標獎金_R", "總計_R", "執行業務_R", "補充保費_R", "應付薪資_R", "年資"
    ]
    
    all_rows = []
    for coach in COACH_PRICING.keys():
        for course in course_types:
            row = {col: 0 for col in columns}
            # 初始化基本資訊
            row["教練_L"], row["課程_L"], row["單價_L"] = coach, course, COACH_PRICING[coach].get(course, 0)
            row["教練_R"], row["課程_R"], row["單價_R"] = coach, course, COACH_PRICING[coach].get(course, 0)
            row["年資"] = SENIORITY_DATA.get(coach, "")
            all_rows.append(row)
    
    df_master = pd.DataFrame(all_rows)
    
    # 讀取上傳檔案並填入數據
    for file in uploaded_files:
        df_info = pd.read_excel(file, sheet_name='統計總表', nrows=1, header=None)
        loc_name = str(df_info.iloc[0, 1]).strip()
        file.seek(0)
        df_stats = pd.read_excel(file, sheet_name='統計總表', skiprows=2).fillna(0)
        
        for _, s_row in df_stats.iterrows():
            raw_name = str(s_row['姓名']).strip().split(' ')[0]
            mapped_name = NAME_CONVERSION.get(raw_name, raw_name)
            
            # 精確定位教練，避免重複統計
            coach_mask = (df_master["教練_L"] == mapped_name)
            
            for course in course_types:
                if course in s_row:
                    val = s_row[course]
                    if val == 0: continue
                    
                    course_mask = (df_master["課程_L"] == course)
                    target_mask = coach_mask & course_mask
                    
                    if loc_name == "巨蛋館":
                        # 僅更新右側巨蛋館欄位
                        df_master.loc[target_mask, "巨蛋館堂數"] = val
                        df_master.loc[target_mask, "巨蛋館金額"] = val * COACH_PRICING.get(mapped_name, {}).get(course, 0)
                    else:
                        # 僅更新左側對應分館欄位
                        s_col, a_col = f"{loc_name}堂數", f"{loc_name}金額"
                        if s_col in df_master.columns:
                            df_master.loc[target_mask, s_col] = val
                            df_master.loc[target_mask, a_col] = val * COACH_PRICING.get(mapped_name, {}).get(course, 0)

    # 計算匯總與稅費
    final_dfs = []
    for coach in COACH_PRICING.keys():
        c_df = df_master[df_master["教練_L"] == coach].copy()
        
        # 三館計算 (左側)
        c_df["三館總堂數"] = c_df[["義昌館堂數", "高美館堂數", "中山館堂數"]].sum(axis=1)
        c_df["三館總金額"] = c_df[["義昌館金額", "高美館金額", "中山館金額"]].sum(axis=1)
        total_L = c_df["三館總金額"].sum()
        tax_L = int(total_L * 0.1) if total_L >= 20000 else 0
        health_L = int(total_L * 0.0211) if total_L >= 20000 else 0
        pay_L = total_L - tax_L - health_L
        
        # 巨蛋館計算 (右側)
        total_R = c_df["巨蛋館金額"].sum()
        tax_R = int(total_R * 0.1) if total_R >= 20000 else 0
        health_R = int(total_R * 0.0211) if total_R >= 20000 else 0
        pay_R = total_R - tax_R - health_R
        
        # 填入匯總欄位
        for col, val in zip(["應付金額_L", "總計_L", "執行業務_L", "補充保費_L", "應付薪資_L"], [total_L, total_L, tax_L, health_L, pay_L]):
            c_df[col] = val
        for col, val in zip(["應付金額_R", "總計_R", "執行業務_R", "補充保費_R", "應付薪資_R"], [total_R, total_R, tax_R, health_R, pay_R]):
            c_df[col] = val
            
        final_dfs.append(c_df)
        
        # 加入小計行
        sub_row = {col: 0 for col in columns}
        sub_row["教練_L"], sub_row["課程_L"], sub_row["應付金額_L"], sub_row["應付薪資_L"] = coach, "小計", total_L, pay_L
        sub_row["教練_R"], sub_row["課程_R"], sub_row["應付金額_R"], sub_row["應付薪資_R"] = coach, "小計", total_R, pay_R
        sub_row["年資"] = SENIORITY_DATA.get(coach, "")
        final_dfs.append(pd.DataFrame([sub_row]))

    df_final = pd.concat(final_dfs, ignore_index=True)
    
    # 輸出 Excel 並排版
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_final.to_excel(writer, index=False, sheet_name='薪資清單')
        ws = writer.sheets['薪資清單']
        
        # 修正標題顯示
        display_headers = [
            "教練", "課程", "單價", "義昌館堂數", "義昌館金額", "高美館堂數", "高美館金額", "中山館堂數", "中山館金額", "三館總堂數", "三館總金額", "應付金額", "堂數達標獎金(加項)", "總計", "執行業務(扣款)", "補充保費(扣款)", "應付薪資",
            "教練", "課程", "單價", "巨蛋館堂數", "巨蛋館金額", "應付金額", "堂數達標獎金(加項)", "總計", "執行業務(扣款)", "補充保費(扣款)", "應付薪資", "年資"
        ]
        for i, h in enumerate(display_headers):
            ws.cell(row=1, column=i+1).value = h
            
        thin = Side(style='thin')
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        center = Alignment(horizontal='center', vertical='center')
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        sub_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        for cell in ws[1]:
            cell.fill, cell.font, cell.alignment, cell.border = header_fill, Font(bold=True), center, border
            
        start_r = 2
        merge_cols = [1, 12, 13, 14, 15, 16, 17, 18, 23, 24, 25, 26, 27, 28, 29]
        
        for r in range(2, ws.max_row + 1):
            coach_val = ws.cell(row=r, column=1).value
            next_coach = ws.cell(row=r + 1, column=1).value if r < ws.max_row else None
            is_sub = ws.cell(row=r, column=2).value == "小計"
            
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                cell.border, cell.alignment = border, center
                if is_sub: cell.fill = sub_fill
            
            if next_coach != coach_val:
                # 執行合併邏輯
                for c in merge_cols:
                    ws.merge_cells(start_row=start_r, start_column=c, end_row=r-1, end_column=c)
                start_r = r + 1
                
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 12
            
    return output.getvalue()

st.title("教練薪資結算系統 (並列修正版)")
uploaded_files = st.file_uploader("上傳預約統計表 (.xlsx)", type=["xlsx"], accept_multiple_files=True)
if uploaded_files:
    if st.button("開始計算薪資並排版"):
        final_excel_data = generate_perfect_salary_report(uploaded_files)
        st.download_button(label="下載薪資明細表", data=final_excel_data, file_name="教練薪資明細_並列修正版.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
