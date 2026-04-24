import streamlit as st
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font

def generate_perfect_salary_report(uploaded_files):
    # 嚴格單價表 (完全依照原始數據)
    COACH_PRICING = {
        "林意潔": {"團1-2人": 800, "團3人": 900, "團4人": 900, "團5人": 950, "團6人": 1000, "1對2(1.5hr)": 1275, "1對2": 850, "1對1(1.5hr)": 1275, "1對1": 850},
        "陳秀蓉": {"團1-2人": 800, "團3人": 850, "團4人": 850, "團5人": 900, "團6人": 950, "1對2(1.5hr)": 1200, "1對2": 800, "1對1(1.5hr)": 1200, "1對1": 800},
        "陳怡廷": {"團1-2人": 800, "團3人": 850, "團4人": 900, "團5人": 950, "團6人": 1000, "1對2(1.5hr)": 1275, "1對2": 850, "1對1(1.5hr)": 1275, "1對1": 850},
        "鍾佳蓁 Rita": {"團1-2人": 900, "團3人": 950, "團4人": 1000, "團5人": 1050, "團6人": 1100, "1對2(1.5hr)": 1425, "1對2": 950, "1對1(1.5hr)": 1425, "1對1": 950},
        "黃宛婷": {"團1-2人": 700, "團3人": 750, "團4人": 800, "團5人": 850, "團6人": 900, "1對2(1.5hr)": 1125, "1對2": 750, "1對1(1.5hr)": 1125, "1對1": 750},
        "楊子慧(小在)": {"團1-2人": 650, "團3人": 800, "團4人": 800, "團5人": 850, "團6人": 900, "1對2(1.5hr)": 1050, "1對2": 700, "1對1(1.5hr)": 1050, "1對1": 700},
        "許力尹 LOUIS": {"團1-2人": 600, "團3人": 800, "團4人": 800, "團5人": 800, "團6人": 800, "1對2(1.5hr)": 1200, "1對2": 800, "1對1(1.5hr)": 1200, "1對1": 800},
        "顥顥": {"團1-2人": 0, "團3人": 1000, "團4人": 1000, "團5人": 1100, "團6人": 1100, "1對2(1.5hr)": 1500, "1對2": 1000, "1對1(1.5hr)": 1500, "1對1": 1000},
        "洪睿絃": {"團1-2人": 700, "團3人": 750, "團4人": 800, "團5人": 850, "團6人": 900, "1對2(1.5hr)": 1125, "1對2": 750, "1對1(1.5hr)": 1125, "1對1": 750},
        "紀儒蓁": {"團1-2人": 700, "團3人": 800, "團4人": 800, "團5人": 850, "團6人": 900, "1對2(1.5hr)": 1125, "1對2": 750, "1對1(1.5hr)": 1125, "1對1": 750},
        "李翎瑋": {"團1-2人": 800, "團3人": 850, "團4人": 900, "團5人": 950, "團6人": 1000, "1對2(1.5hr)": 1275, "1對2": 850, "1對1(1.5hr)": 1275, "1對1": 850},
        "郭奕伶": {"團1-2人": 550, "團3人": 600, "團4人": 650, "團5人": 700, "團6人": 750, "1對2(1.5hr)": 900, "1對2": 600, "1對1(1.5hr)": 900, "1對1": 600},
        "郭品均": {"團1-2人": 700, "團3人": 750, "團4人": 800, "團5人": 850, "團6人": 900, "1對2(1.5hr)": 1125, "1對2": 750, "1對1(1.5hr)": 1125, "1對1": 750},
        "邴妍語": {"團1-2人": 700, "團3人": 750, "團4人": 800, "團5人": 850, "團6人": 900, "1對2(1.5hr)": 1125, "1對2": 750, "1對1(1.5hr)": 1125, "1對1": 750},
        "張鈞弼": {"團1-2人": 550, "團3人": 600, "團4人": 650, "團5人": 700, "團6人": 850, "1對2(1.5hr)": 900, "1對2": 600, "1對1(1.5hr)": 900, "1對1": 600},
        "蕭竣升": {"團1-2人": 700, "團3人": 750, "團4人": 800, "團5人": 850, "團6人": 900, "1對2(1.5hr)": 1125, "1對2": 750, "1對1(1.5hr)": 1125, "1對1": 750},
        "紀萃文": {"團1-2人": 700, "團3人": 750, "團4人": 800, "團5人": 850, "團6人": 900, "1對2(1.5hr)": 1125, "1對2": 750, "1對1(1.5hr)": 1125, "1對1": 750},
        "李函豫": {"團1-2人": 600, "團3人": 650, "團4人": 700, "團5人": 750, "團6人": 800, "1對2(1.5hr)": 975, "1對2": 650, "1對1(1.5hr)": 975, "1對1": 650},
        "尤子綺": {"團1-2人": 550, "團3人": 600, "團4人": 650, "團5人": 700, "團6人": 750, "1對2(1.5hr)": 900, "1對2": 600, "1對1(1.5hr)": 900, "1對1": 600},
        "張楷翌": {"團1-2人": 600, "團3人": 650, "團4人": 700, "團5人": 750, "團6人": 800, "1對2(1.5hr)": 975, "1對2": 650, "1對1(1.5hr)": 975, "1對1": 650},
        "侯懿庭": {"團1-2人": 700, "團3人": 750, "團4人": 800, "團5人": 850, "團6人": 900, "1對2(1.5hr)": 1125, "1對2": 750, "1對1(1.5hr)": 1125, "1對1": 750},
        "謝俐池": {"團1-2人": 550, "團3人": 600, "團4人": 650, "團5人": 700, "團6人": 750, "1對2(1.5hr)": 900, "1對2": 600, "1對1(1.5hr)": 900, "1對1": 600},
        "黃姿菁": {"團1-2人": 550, "團3人": 600, "團4人": 650, "團5人": 700, "團6人": 750, "1對2(1.5hr)": 900, "1對2": 600, "1對1(1.5hr)": 900, "1對1": 600},
        "籃郁雯": {"團1-2人": 550, "團3人": 600, "團4人": 700, "團5人": 700, "團6人": 700, "1對2(1.5hr)": 900, "1對2": 600, "1對1(1.5hr)": 900, "1對1": 600},
        "徐漫": {"團1-2人": 500, "團3人": 550, "團4人": 600, "團5人": 650, "團6人": 700, "1對2(1.5hr)": 825, "1對2": 550, "1對1(1.5hr)": 825, "1對1": 550},
        "鄭筠馨": {"團1-2人": 550, "團3人": 600, "團4人": 650, "團5人": 700, "團6人": 750, "1對2(1.5hr)": 900, "1對2": 600, "1對1(1.5hr)": 900, "1對1": 600},
        "高舒涵": {"團1-2人": 550, "團3人": 600, "團4人": 650, "團5人": 700, "團6人": 750, "1對2(1.5hr)": 900, "1對2": 600, "1對1(1.5hr)": 900, "1對1": 600},
        "邱靜瑜": {"團1-2人": 600, "團3人": 650, "團4人": 700, "團5人": 750, "團6人": 800, "1對2(1.5hr)": 975, "1對2": 650, "1對1(1.5hr)": 975, "1對1": 650}
    }

    NAME_CONVERSION = {
        "意潔": "林意潔", "Vivi": "陳秀蓉", "怡廷": "陳怡廷", "佳蓁": "鍾佳蓁 Rita", "宛婷": "黃宛婷",
        "小在": "楊子慧(小在)", "LOUIS": "許力尹 LOUIS", "顥顥": "顥顥", "睿絃": "洪睿絃", "儒蓁": "紀儒蓁",
        "翎瑋": "李翎瑋", "奕伶": "郭奕伶", "品均": "郭品均", "妍語": "邴妍語", "鈞弼": "張鈞弼",
        "竣升": "蕭竣升", "萃萃": "紀萃文", "函豫": "李函豫", "子綺": "尤子綺", "楷翌": "張楷翌",
        "懿庭": "侯懿庭", "俐池": "謝俐池", "姿菁": "黃姿菁", "郁雯": "籃郁雯", "徐漫": "徐漫",
        "筠馨": "鄭筠馨", "舒涵": "高舒涵", "靜瑜": "邱靜瑜"
    }

    course_types = ["團1-2人", "團3人", "團4人", "團5人", "團6人", "1對2(1.5hr)", "1對2", "1對1(1.5hr)", "1對1"]

    columns = [
        "教練", "課程", "單價",
        "義昌館堂數", "義昌館金額", "高美館堂數", "高美館金額", "中山館堂數", "中山館金額",
        "堂數達標獎金", "三館總堂數", "三館總金額",
        "應付金額", "堂數達標獎金(加項)", "總計", "執行業務(扣款)", "補充保費(扣款)", "應付薪資"
    ]

    # 初始化數據容器
    all_rows = []
    for coach in COACH_PRICING.keys():
        for course in course_types:
            row = {col: 0 for col in columns}
            row["教練"] = coach
            row["課程"] = course
            row["單價"] = COACH_PRICING[coach].get(course, 0)
            all_rows.append(row)
    df_master = pd.DataFrame(all_rows)

    # 讀取上傳的檔案
    for file in uploaded_files:
        try:
            # 讀取第一列獲取館別
            df_info = pd.read_excel(file, sheet_name='統計總表', nrows=1, header=None)
            loc_name = str(df_info.iloc[0, 1]).strip()
            
            if loc_name not in ["義昌館", "高美館", "中山館"]:
                st.warning(f"⚠️ 檔案 {file.name} 的館別 [{loc_name}] 不在預期內，已略過此檔案。")
                continue

            # 重新將檔案指標移回開頭，以供後續完整讀取
            file.seek(0)
            df_stats = pd.read_excel(file, sheet_name='統計總表', skiprows=2).fillna(0)
            
            if '團1人' in df_stats.columns and '團2人' in df_stats.columns:
                df_stats['團1-2人'] = df_stats['團1人'] + df_stats['團2人']

            s_col, a_col = f"{loc_name}堂數", f"{loc_name}金額"
            for idx, m_row in df_master.iterrows():
                for _, s_row in df_stats.iterrows():
                    s_name = str(s_row['姓名']).strip().split(' ')[0]
                    if NAME_CONVERSION.get(s_name, s_name) == m_row['教練']:
                        if m_row['課程'] in s_row:
                            val = s_row[m_row['課程']]
                            df_master.at[idx, s_col] = val
                            df_master.at[idx, a_col] = val * m_row['單價']
                        break
        except Exception as e:
            st.error(f"❌ 讀取檔案 {file.name} 時發生錯誤: {e}")

    # 計算總額與扣款
    final_dfs = []
    for coach in COACH_PRICING.keys():
        c_df = df_master[df_master["教練"] == coach].copy()

        c_df["三館總堂數"] = c_df[["義昌館堂數", "高美館堂數", "中山館堂數"]].sum(axis=1)
        c_df["三館總金額"] = c_df[["義昌館金額", "高美館金額", "中山館金額"]].sum(axis=1)

        base_pay = c_df["三館總金額"].sum()
        bonus_plus = 0  # 堂數達標獎金(加項) - 先預設為 0
        total_sum = base_pay + bonus_plus

        if total_sum >= 20000:
            tax_10 = int(total_sum * 0.1)
            health_211 = int(total_sum * 0.0211)
        else:
            tax_10 = 0
            health_211 = 0

        final_pay = total_sum - tax_10 - health_211

        c_df["應付金額"] = base_pay
        c_df["堂數達標獎金(加項)"] = bonus_plus
        c_df["總計"] = total_sum
        c_df["執行業務(扣款)"] = tax_10
        c_df["補充保費(扣款)"] = health_211
        c_df["應付薪資"] = final_pay

        final_dfs.append(c_df)

        sub_row = {col: 0 for col in columns}
        sub_row["教練"] = coach
        sub_row["課程"] = "小計"
        for col in ["義昌館堂數", "義昌館金額", "高美館堂數", "高美館金額", "中山館堂數", "中山館金額", "三館總堂數", "三館總金額"]:
            sub_row[col] = c_df[col].sum()
        sub_row["應付金額"] = base_pay
        sub_row["總計"] = total_sum
        sub_row["應付薪資"] = final_pay
        final_dfs.append(pd.DataFrame([sub_row]))

    df_final = pd.concat(final_dfs, ignore_index=True)

    # 寫入記憶體中的 BytesIO 物件
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_final.to_excel(writer, index=False, sheet_name='薪資清單')
        ws = writer.sheets['薪資清單']

        thin = Side(style='thin')
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        center = Alignment(horizontal='center', vertical='center')
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        sub_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True)
            cell.alignment = center
            cell.border = border

        start_r = 2
        pay_idx = df_final.columns.get_loc("應付金額") + 1  

        for r in range(2, ws.max_row + 1):
            coach_val = ws.cell(row=r, column=1).value
            next_coach = ws.cell(row=r + 1, column=1).value if r < ws.max_row else None

            for c in range(1, ws.max_column + 1):
                ws.cell(row=r, column=c).border = border
                ws.cell(row=r, column=c).alignment = center
                if ws.cell(row=r, column=2).value == "小計":
                    ws.cell(row=r, column=c).fill = sub_fill

            if next_coach != coach_val:
                ws.merge_cells(start_row=start_r, start_column=1, end_row=r, end_column=1)
                for c in range(pay_idx, ws.max_column + 1):
                    ws.merge_cells(start_row=start_r, start_column=c, end_row=r, end_column=c)
                start_r = r + 1

        for col in ws.columns:
            max_l = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = max_l + 2

    # 回傳二進位資料
    processed_data = output.getvalue()
    return processed_data

# ==========================================
# Streamlit UI 介面設計
# ==========================================
st.set_page_config(page_title="教練薪資結算系統", layout="centered")

st.title("💰 教練薪資結算系統")
st.markdown("請將當月的 **義昌館、高美館、中山館** 統計表 Excel 檔上傳。您可以一次框選多個檔案拖曳進來。")

# 1. 檔案上傳區塊
uploaded_files = st.file_uploader("上傳預約統計表 (.xlsx)", type=["xlsx"], accept_multiple_files=True)

# 2. 處理與下載邏輯
if uploaded_files:
    # 顯示目前上傳的檔案清單
    st.success(f"已成功上傳 {len(uploaded_files)} 個檔案！")
    
    # 點擊按鈕後開始處理
    if st.button("🚀 開始計算薪資並排版"):
        with st.spinner("系統正在處理數據與 Excel 排版，請稍候..."):
            
            # 呼叫主函式取得處理好的 Excel 二進位檔案
            final_excel_data = generate_perfect_salary_report(uploaded_files)
            
            st.balloons()
            st.success("✅ 計算與排版已完成！點擊下方按鈕下載最終報表。")
            
            # 3. 下載按鈕
            st.download_button(
                label="📥 下載薪資明細表",
                data=final_excel_data,
                file_name="教練薪資明細_最終版.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
else:
    st.info("等待上傳檔案中...")